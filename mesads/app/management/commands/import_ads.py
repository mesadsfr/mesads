import argparse
import datetime
import functools
import itertools
import re
import string
import sys
from unidecode import unidecode

import openpyxl

from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand
from django.db import transaction

from reversion.revisions import create_revision, set_user
from mesads.users.models import User

from mesads.app.models import ADS, ADSManager, ADSUser, validate_siret
from mesads.fradm.models import Commune, Prefecture


class Excel:
    """Helper to manipulate an excel file"""

    def __init__(self, header):
        self.header = header
        self.normalized_header = [h.lower() for h in self.header]

    def colnames(self):
        """Return all columns names in order (A, B, C, ..., Z, AA, AB, AC, ...,
        AZ, BA, BB, BC, ..., ZZ, AAA, AAB, ... ZZZ)"""
        return (
            "".join(v)
            for v in itertools.chain(
                itertools.product(string.ascii_uppercase, repeat=1),
                itertools.product(string.ascii_uppercase, repeat=2),
                itertools.product(string.ascii_uppercase, repeat=3),
            )
        )

    def idx_to_colname(self, nth):
        """Return the nth excel column name"""
        return next(itertools.islice(self.colnames(), nth, nth + 1))

    def find_cols(self, colname, exact=False):
        """Return the index of the given column name. If exact=False, the column
        name can be partial, but must be unambiguous."""
        candidates = []
        last_match = None
        for idx, col in enumerate(self.normalized_header):
            # Exact match in one of the lines of the header, or partial match
            if (exact and colname.lower() in col.split("\n")) or (
                not exact and colname.lower() in col
            ):
                candidates.append(idx)
                if last_match and last_match.lower() != col.lower():
                    err_last_match = "\n".join(
                        f"\t{line}" for line in last_match.splitlines()
                    )
                    err_col = "\n".join(f"\t{line}" for line in col.splitlines())

                    raise ValueError(
                        f'La colonne "{colname}" est trouvée plusieurs fois avec des valeurs différentes, par exemple :\n{err_last_match}\net :\n{err_col}. Régler le problème ou ajouter exact=True'
                    )
                last_match = col

        if len(candidates) == 0:
            raise ValueError(f'Colonne "{colname}" inconnue')
        return candidates

    def idx(self, colname, exact=False):
        cols = self.find_cols(colname, exact=exact)
        if len(cols) > 1:
            raise ValueError(
                f"Colonne \"{colname}\" ambiguë ({len(cols)} résultats: {', '.join((self.idx_to_colname(candidate) for candidate in cols))})"
            )
        return cols[0]

    def nidx(self, colname, nth, exact=False):
        """Return the index of the nth column with the given name"""
        cols = self.find_cols(colname, exact=exact)
        return cols[nth]


class ADSImporter:
    COLUMNS = (
        # Col 0: A
        "Numéro du département de la commune",
        # Col 1: B
        "Nom de la commune",
        # Col 2: C
        "Numéro de l'ADS",
        # Col 3: D
        'ADS actuellement exploitée ?\n\nMettre "oui" ou "non". Ce champ est obligatoire.',
        # Col 4: E
        "Date de création de l'ADS\n\ndate à laquelle l'ADS a été attribuée pour la première fois",
        # Col 5: F
        "Date du dernier renouvellement de l'ADS\nRemplir seulement si l'ADS a été attribuée pour la première fois après le 01/10/2014 : Les ADS créées depuis le 1er octobre 2014 sont valables 5 ans et doivent être renouvelées.",
        # Col 6: G
        "Date d'attribution de l'ADS au titulaire actuel\nLaissez ce champ vide si le titulaire n'a pas changé depuis la création de l'ADS. Ne remplir que si l'ADS a été attribuée pour la première fois avant le 01/10/2014",
        # Col 7: H
        'Véhicule conventionné CPAM ?\nMettre "oui" ou "non". Laissez vide si vous ne savez pas.',
        # Col 8: I
        "Plaque d'immatriculation du véhicule",
        # Col 9: J
        'Le véhicule est-il un véhicule électrique/hybride ?\nMettre "oui" ou "non", laissez vide si vous ne savez pas',
        # Col 10: K
        'Véhicule compatible PMR ?\nMettre "oui" ou "non", laissez vide si vous ne savez pas',
        # Col 11: L
        "Titulaire de l'ADS\nS'il s'agit d'une personne physique, précisez le nom et le prénom du titulaire de l'ADS. S'il s'agit d'une personne morale, indiquez sa raison sociale.",
        # Col 12: M
        "SIRET du titulaire de l'ADS",
        # Col 13: N
        "Téléphone fixe du titulaire de l'ADS",
        # Col 14: O
        "Téléphone mobile du titulaire de l'ADS",
        # Col 15: P
        "Email du titulaire de l'ADS",
        #
        # 1er conducteur
        #
        # Col 16: Q
        "Entrez dans les 4 colonnes à droite les informations du 1e conducteur",
        # Col 17: R
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 18: S
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 19: T
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 20: U
        "Numéro de la carte professionnelle",
        #
        # 2e conducteur
        #
        # Col 21: V
        "Entrez dans les 4 colonnes à droite les informations du 2e conducteur",
        # Col 22: W
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 23: X
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 24: Y
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 25: Z
        "Numéro de la carte professionnelle",
        #
        # 3e conducteur
        #
        # Col 26: AA
        "Entrez dans les 4 colonnes à droite les informations du 3e conducteur",
        # Col 27: AB
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 28: AC
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 29: AD
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 30: AE
        "Numéro de la carte professionnelle",
        #
        # 4e conducteur
        #
        # Col 31: AF
        "Entrez dans les 4 colonnes à droite les informations du 4e conducteur",
        # Col 32: AG
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 33: AH
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 34: AI
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 35: AJ
        "Numéro de la carte professionnelle",
        #
        # 5e conducteur
        #
        # Col 35: AK
        "Entrez dans les 4 colonnes à droite les informations du 5e conducteur",
        # Col 36: AL
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 37: AM
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 38: AN
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 30: AE
        "Numéro de la carte professionnelle",
        #
        # 6e conducteur
        #
        # Col 31: AF
        "Entrez dans les 4 colonnes à droite les informations du 6e conducteur",
        # Col 32: AG
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 33: AH
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 34: AI
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 35: AJ
        "Numéro de la carte professionnelle",
        #
        # 7e conducteur
        #
        # Col 36: AK
        "Entrez dans les 4 colonnes à droite les informations du 7e conducteur",
        # Col 37: AL
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 38: AM
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 39: AN
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 40: AO
        "Numéro de la carte professionnelle",
        #
        # 8e conducteur
        #
        # Col 41: AP
        "Entrez dans les 4 colonnes à droite les informations du 8e conducteur",
        # Col 42: AQ
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 43: AR
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 44: AS
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 45: AT
        "Numéro de la carte professionnelle",
        #
        # 9e conducteur
        #
        # Col 46: AU
        "Entrez dans les 4 colonnes à droite les informations du 9e conducteur",
        # Col 47: AV
        "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :\n"
        "TITULAIRE si le conducteur est le titulaire de l'ADS (personne physique). Dans ce cas, renseignez uniquement le numéro de la carte professionnelle.\n"
        "REPRESENTANT si le conducteur est le représentant légal de la société titulaire de l'ADS (gérant ou président non salarié)\n"
        "SALARIE si le conducteur est salarié du titulaire de l'ADS\n"
        "LOCATAIRE COOPERATEUR si le conducteur est locataire coopérateur\n"
        "LOCATAIRE GERANT si le conducteur est locataire gérant",
        # Col 48: AW
        "Nom du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 49: AX
        "SIRET du conducteur de l'ADS\n"
        "À remplir uniquement si le conducteur n'est pas TITULAIRE",
        # Col 50: AY
        "Numéro de la carte professionnelle",
    )

    def __init__(self):
        self.excel = Excel(self.COLUMNS)

    def check_header(self, cols):
        """Make sure the header has the columns we expect. We ignore minimal differences (case, spaces, etc.)"""
        for idx, (col, exp) in enumerate(itertools.zip_longest(cols, self.COLUMNS)):
            sanitized_col = (
                (col or "").strip().lower().replace("\n", "").replace(" ", "")
            )
            sanitized_exp = (
                (exp or "").strip().lower().replace("\n", "").replace(" ", "")
            )

            if sanitized_col == sanitized_exp:
                continue
            raise self.fmt_col_error(
                f"valeur attendue\nvvvvvvvvvvvvvvvvvvvvvvvv\n{exp}\n^^^^^^^^^^^^^^^^^^^^^^^^\n",
                col,
                idx,
            )

    def fmt_col_error(self, msg, value, idx):
        return ValueError(f"{msg}. Colonne {self.excel.idx_to_colname(idx)}: {value}")

    def load_ads(self, cols, override=False):
        departement = self.find_departement(
            cols[self.excel.idx("numéro du département")]
        )
        commune = self.find_commune(
            departement,
            cols[self.excel.idx("nom de la commune")],
        )
        ads_manager = self.find_ads_manager(commune)
        ads_number = cols[
            self.excel.idx(
                "numéro de l'ads",
                exact=True,
            )
        ]
        ads = self.find_existing_ads(ads_manager, ads_number)
        if not ads:
            ads = ADS(ads_manager=ads_manager, number=ads_number)
        elif not override:
            raise self.fmt_col_error(
                f"ADS {ads_number} déjà existante, mais le paramètre --override n'a pas été spécifié",
                f"ADS id={ads.id}",
                self.excel.idx("numéro de l'ads", exact=True),
            )

        ads.ads_in_use = self.parse_bool(
            cols,
            self.excel.idx("ADS actuellement exploitée ?"),
        )

        ads.ads_creation_date = self.parse_date(
            cols,
            self.excel.idx("date de création"),
        )
        ads.ads_renew_date = self.parse_date(
            cols,
            self.excel.idx("date du dernier renouvellement"),
        )
        ads.attribution_date = self.parse_date(
            cols,
            self.excel.idx("date d'attribution"),
        )
        ads.accepted_cpam = self.parse_bool(
            cols,
            self.excel.idx("véhicule conventionné CPAM"),
        )
        ads.immatriculation_plate = self.parse_license_plate(
            cols,
            self.excel.idx(
                "immatriculation du véhicule",
            ),
        )
        ads.vehicle_compatible_pmr = self.parse_bool(
            cols,
            self.excel.idx("compatible PMR"),
        )
        ads.owner_name = self.parse_owner_name(
            cols,
            self.excel.idx(
                "titulaire de l'ads",
                exact=True,
            ),
        )
        ads.owner_siret = self.parse_siret(cols, self.excel.idx("siret du titulaire"))
        ads.owner_phone = self.parse_phone(
            cols, self.excel.idx("téléphone fixe du titulaire"), mobile=False
        )
        ads.owner_mobile = self.parse_phone(
            cols, self.excel.idx("téléphone mobile du titulaire"), mobile=True
        )
        ads.owner_email = self.parse_email(
            cols,
            self.excel.idx("email du titulaire"),
        )

        ads_users = self.load_ads_users(cols, ads)
        return ads, ads_users

    def load_ads_users(self, cols, ads):
        """Load the ADS users from the excel file"""
        ads_users = []
        for i in range(9):
            status_exploitant_idx = self.excel.nidx(
                "Qui est le conducteur du véhicule ? Remplir avec une des valeurs suivantes :",
                i,
            )
            date_creation_idx = self.excel.idx("date de création")
            name_exploitant_idx = self.excel.nidx("nom du conducteur", i)
            siret_exploitant_idx = self.excel.nidx("siret du conducteur", i)
            license_number_exploitant_idx = self.excel.nidx(
                "numéro de la carte professionnelle", i, exact=True
            )
            if (
                self.is_empty(cols[status_exploitant_idx])
                and self.is_empty(cols[name_exploitant_idx])
                and self.is_empty(cols[siret_exploitant_idx])
                and self.is_empty(cols[license_number_exploitant_idx])
            ):
                continue

            # Content of the first non-empty column set between name, siret and license number. Used to report errors.
            first_col_set = (
                cols[status_exploitant_idx]
                or cols[name_exploitant_idx]
                or cols[siret_exploitant_idx]
                or (cols[license_number_exploitant_idx])
            )
            # Index of the first non-empty column. Used to report errors.
            first_col_set_idx = (
                (cols[status_exploitant_idx] and status_exploitant_idx)
                or (cols[name_exploitant_idx] and name_exploitant_idx)
                or (cols[siret_exploitant_idx] and siret_exploitant_idx)
                or (
                    cols[license_number_exploitant_idx]
                    and license_number_exploitant_idx
                )
            )

            if not ads.ads_creation_date:
                raise self.fmt_col_error(
                    "La date de création de l'ADS doit être renseignée avant de renseigner les exploitants",
                    ads.ads_creation_date,
                    date_creation_idx,
                )

            if ads.ads_creation_date >= datetime.date(2014, 10, 1) and i > 1:
                raise self.fmt_col_error(
                    "Un seul exploitant peut être défini pour les nouvelles ADS",
                    first_col_set,
                    first_col_set_idx,
                )

            status = self.parse_exploitant_status(cols, status_exploitant_idx)

            name = cols[name_exploitant_idx] or ""
            if name and status == "titulaire_exploitant":
                if str(name).replace(" ", "") == ads.owner_name.replace(" ", ""):
                    name = ""
                else:
                    raise self.fmt_col_error(
                        'Pour les titulaires exploitants, le nom doit être vide et rempli dans le champ "Titulaire de l\'ADS"',
                        first_col_set,
                        first_col_set_idx,
                    )

            siret = cols[siret_exploitant_idx] or ""
            if siret and (
                status == "titulaire_exploitant" or status == "legal_representative"
            ):
                if siret == ads.owner_siret:
                    siret = ""
                else:
                    raise self.fmt_col_error(
                        'Pour les titulaires exploitants et les représentants légaux, le SIRET doit être vide et rempli dans le champ "SIRET du titulaire de l\'ADS"',
                        first_col_set,
                        first_col_set_idx,
                    )

            license_number = cols[license_number_exploitant_idx] or ""

            ads_users.append(
                ADSUser(
                    ads=ads,
                    status=status,
                    name=name,
                    siret=siret,
                    license_number=license_number,
                )
            )
        return ads_users

    @functools.cache
    def find_departement(self, numero):
        if not numero:
            raise ValueError("Le département doit être renseigné")

        if numero < 10:
            numero = "0" + str(numero)

        query = Prefecture.objects.filter(numero=numero)
        res = query.all()
        if len(res) == 0:
            raise ValueError(f"Département {numero} inconnu")
        if len(res) > 1:
            raise ValueError(f"Département {numero} ambigu ({len(res)} résultats)")
        return res[0]

    def normalize(self, value):
        """Remove all the special chars of `value`. Useful to compare two names."""
        return "".join(
            c
            for c in unidecode(value)
            if c in string.ascii_letters or c in string.digits
        ).lower()

    @functools.cache
    def find_commune(self, departement, name):
        query = Commune.objects.filter(
            departement=departement.numero, libelle__iexact=name
        )
        res = query.all()
        if len(res) == 1:
            return res[0]

        query = Commune.objects.filter(departement=departement.numero)
        normalized_name = self.normalize(name)
        candidates = []
        for commune in query.all():
            if self.normalize(commune.libelle) == normalized_name:
                candidates.append(commune)

        if len(candidates) == 0:
            raise ValueError(
                f"Commune {name} inconnue dans le département {departement}"
            )
        elif len(candidates) > 1:
            raise ValueError(
                f"Commune {name} ambiguë dans le département {departement}. Résultats trouvés: {', '.join((candidate.libelle for candidate in candidates))}"
            )
        return candidates[0]

    @functools.cache
    def find_ads_manager(self, commune):
        return ADSManager.objects.filter(
            content_type=ContentType.objects.get_for_model(Commune),
            object_id=commune.id,
        ).get()

    def find_existing_ads(self, ads_manager, numero):
        return ADS.objects.filter(ads_manager=ads_manager, number=numero).first()

    def parse_date(self, cols, idx):
        """Convert datetime to a value"""
        value = cols[idx]
        if not value:
            return None
        if not isinstance(value, datetime.datetime):
            raise self.fmt_col_error("Date invalide", value, idx)
        return value.date()

    def parse_bool(self, cols, idx):
        """Convert a string to a boolean value"""
        value = cols[idx]
        if not value:
            return None
        if value.lower().strip() == "oui":
            return True
        elif value.lower().strip() == "non":
            return False
        raise self.fmt_col_error("Valeur invalide", value, idx)

    def parse_license_plate(self, cols, idx):
        return cols[idx] or ""

    def parse_owner_name(self, cols, idx):
        return cols[idx] or ""

    def parse_siret(self, cols, idx):
        siret = "".join(str(cols[idx]).strip().split())
        if siret == "None":
            return ""
        if siret:
            try:
                validate_siret(siret)
            except Exception:
                raise self.fmt_col_error("Siret invalide", siret, idx)
        return siret

    def parse_phone(self, cols, idx, mobile=False):
        if not cols[idx]:
            return ""
        phone = "".join(str(cols[idx]).replace(".", "").split())
        # Accept 9 digits phone numbers
        if len(phone) == 9:
            phone = "0" + phone

        if len(phone) != 10:
            raise self.fmt_col_error(
                "Numéro de téléphone invalide, doit faire 10 chiffres", cols[idx], idx
            )
        if not phone.startswith("0"):
            raise self.fmt_col_error(
                "Numéro de téléphone invalide, doit commencer par 0", cols[idx], idx
            )
        if mobile and phone:
            if not phone.startswith("06") and not phone.startswith("07"):
                raise self.fmt_col_error(
                    "Numéro de téléphone invalide, doit commencer par 06 ou 07",
                    cols[idx],
                    idx,
                )
        elif not mobile and phone:
            if phone.startswith("06") or phone.startswith("07"):
                raise self.fmt_col_error(
                    "Numéro de téléphone invalide, ne doit pas commencer par 06 ou 07",
                    cols[idx],
                    idx,
                )
        return phone

    def parse_email(self, cols, idx):
        if not cols[idx]:
            return ""
        if not re.match(r"^.*@.*\..*$", cols[idx]):
            raise self.fmt_col_error("Email invalide", cols[idx], idx)
        return cols[idx]

    def parse_license_number(self, cols, idx):
        return cols[idx] or ""

    def parse_exploitant_status(self, cols, idx):
        choices = {
            "titulaire": "titulaire_exploitant",
            "représentant": "legal_representative",
            "salarié": "salarie",
            "locataire coopérateur": "cooperateur",
            "locataire gérant": "locataire_gerant",
        }
        ret = choices.get((cols[idx] or "").lower().strip())
        if not ret:
            raise self.fmt_col_error("Statut de l'exploitant invalide", cols[idx], idx)
        return ret

    def is_empty(self, value):
        if value is None:
            return True
        return str(value).strip() == ""


class Command(BaseCommand):
    help = "Import ADS from CSV file"

    def add_arguments(self, parser):
        parser.add_argument("-f", "--ads-file", required=True)
        parser.add_argument(
            "--save", action=argparse.BooleanOptionalAction, default=False
        )
        parser.add_argument(
            "--override", action=argparse.BooleanOptionalAction, default=False
        )

    def _log(self, level, msg, icon=None):
        if not icon:
            icon = (level == self.style.SUCCESS) and "✅" or "❌"
        sys.stdout.write(level(f"{icon} {msg}\n"))

    def handle(self, ads_file, override=False, save=False, **opts):
        workbook = openpyxl.load_workbook(ads_file)
        sheet = workbook.active
        importer = ADSImporter()

        header = next(sheet.iter_rows(values_only=True))
        importer.check_header(header)

        has_error = False
        ads_list = []

        self._log(self.style.SUCCESS, "Préparation de l'import...")

        for idx, cols in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            # All columns are empty
            if len([v for v in cols if v is not None and str(v).strip() != ""]) == 0:
                continue

            try:
                ads, ads_users = importer.load_ads(cols, override=override)
                if ads.id:
                    self._log(
                        self.style.SUCCESS,
                        f"Ligne {idx+2}: ADS {ads.id} prête à être mise à jour",
                    )
                else:
                    self._log(
                        self.style.SUCCESS,
                        f"Ligne {idx+2}: préparation d'une nouvelle ADS",
                    )
                ads_list.append(
                    {
                        "ads": ads,
                        "ads_users": ads_users,
                    }
                )
            except ValueError as exc:
                self._log(self.style.ERROR, f"Line {idx+2}: {exc}")
                has_error = True

        if has_error:
            self._log(
                self.style.ERROR,
                "Échec de la préparation de l'import. Réglez les erreurs et réessayez.",
            )
            return

        self._log(
            self.style.SUCCESS,
            "Préparation de l'import terminée. Enregistrement des ADS...",
        )

        # Save all ADS in a single transaction, so that if one fails, all are rolled back
        # create_revision and set_user are called to make sure a
        # reversion_revision entry is created for this initial import.
        try:
            last_exc = None
            with (
                transaction.atomic(),
                create_revision(),
            ):
                revision_user = User.objects.get(email="julien.castets@beta.gouv.fr")
                set_user(revision_user)

                for idx, ads in enumerate(ads_list):
                    try:
                        # For each ADS, we create a new transaction. If one
                        # fails, we still try to save the others but the outer
                        # transaction will be rolled back.
                        with transaction.atomic():
                            ads["ads"].save()
                            self._log(
                                self.style.SUCCESS,
                                f"Ligne {idx+1}: ADS numéro {ads['ads'].number} enregistrée",
                            )

                            for ads_user in ads["ads_users"]:
                                ads_user.save()
                                self._log(
                                    self.style.SUCCESS,
                                    f"Exploitant {ads_user.name} /  {ads_user.siret} / {ads_user.license_number} enregistré",
                                    icon="  ",
                                )

                    except Exception as exc:
                        last_exc = exc
                        self._log(
                            self.style.ERROR,
                            f"Ligne {idx+2}: échec de l'import de l'ADS {ads['ads'].number}: {exc}",
                        )

                if last_exc:
                    raise last_exc
                elif not save:
                    self._log(
                        self.style.WARNING,
                        "Les ADS ne sont pas enregistrées car le paramètre --save n'a pas été spécifié",
                    )
                    raise ValueError
        except:  # noqa
            self._log(
                self.style.ERROR,
                "Échec de l'import, aucune ADS n'a été enregistrée",
            )
        else:
            self._log(self.style.SUCCESS, "Enregistrement terminé.")
